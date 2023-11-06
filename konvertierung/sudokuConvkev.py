#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
import os
import pandas as pd

def generiere_dateiname(dateipfad):
    zaehler = 1
    neuer_dateiname = dateipfad

    while os.path.exists(neuer_dateiname):
        zaehler += 1
        basisname, erweiterung = os.path.splitext(dateipfad)
        neuer_dateiname = f"{basisname}{zaehler}{erweiterung}"

    return neuer_dateiname

def datei_lesen(dateipfad):
    with open(dateipfad, 'r') as datei:
        zeilen = datei.readlines()
    matrix = []
    for zeile in zeilen:
        werte = zeile.strip().split()
        matrix.append(werte)
    return matrix



def bearbeite_matrix_und_speichere_als_excel(matrix, ausgabedatei):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_index, row_data in enumerate(matrix):
        for col_index, cell_value in enumerate(row_data):
            cell = ws.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
            if cell_value == '0':
                cell.value = ' '            
            cell.font = Font(size=26)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    cell_width_cm = 2
    cell_height_cm = 1

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 3.5) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows():
        row_height = 30 
        ws.row_dimensions[row[0].row].height = row_height
    border = Border(left=Side(border_style='thin'), 
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    for i in range(0, 9, 3):
        for j in range(0, 9, 3):
            for x in range(3):
                for y in range(3):
                    cell = ws.cell(row=i + x + 1, column=j + y + 1)
                    cell.border = Border(left=Side(style='medium' if y == 0 else 'thin'),
                                         right=Side(style='medium' if y == 2 else 'thin'),
                                         top=Side(style='medium' if x == 0 else 'thin'),
                                         bottom=Side(style='medium' if x == 2 else 'thin'))
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins.left = 0.5  
    ws.page_margins.right = 0.5  
    ws.page_margins.top = 0.5  
    ws.page_margins.bottom = 0.5  
    ws.print_area = ws.dimensions
    wb.save(ausgabedatei)
    
if __name__ == "__main__":
    dateipfade = ["sudoku1.txt", "sudoku2.txt", "sudoku3.txt", "sudoku4.txt", "sudoku5.txt",
                  "sudoku6.txt", "sudoku7.txt", "sudoku8.txt", "sudoku9.txt", "sudoku10.txt",
                  "sudoku11.txt", "sudoku12.txt", "sudoku13.txt", "sudoku14.txt", "sudoku15.txt",
                  "sudoku16.txt", "sudoku17.txt", "sudoku18.txt", "sudoku19.txt", "sudoku20.txt",
                  "sudoku21.txt", "sudoku22.txt", "sudoku23.txt", "sudoku24.txt", "sudoku25.txt",
                  "sudoku26.txt", "sudoku27.txt", "sudoku28.txt", "sudoku29.txt", "sudoku30.txt",
                  "sudoku31.txt", "sudoku32.txt", "sudoku33.txt", "sudoku34.txt", "sudoku35.txt",
                  "sudoku36.txt", "sudoku37.txt", "sudoku38.txt", "sudoku39.txt", "sudoku40.txt",
                  "sudoku41.txt", "sudoku42.txt", "sudoku43.txt", "sudoku44.txt", "sudoku45.txt",
                  "sudoku46.txt", "sudoku47.txt", "sudoku48.txt", "sudoku49.txt", "sudoku50.txt",
                  "sudoku51.txt", "sudoku52.txt", "sudoku53.txt", "sudoku54.txt", "sudoku55.txt",
                  "sudoku56.txt", "sudoku57.txt", "sudoku58.txt", "sudoku59.txt", "sudoku60.txt",
                  "sudoku61.txt", "sudoku62.txt", "sudoku63.txt", "sudoku64.txt", "sudoku65.txt",
                  "sudoku66.txt", "sudoku67.txt", "sudoku68.txt", "sudoku69.txt", "sudoku70.txt",
                  "sudoku71.txt", "sudoku72.txt", "sudoku73.txt", "sudoku74.txt", "sudoku75.txt",
                  "sudoku76.txt", "sudoku77.txt", "sudoku78.txt", "sudoku79.txt", "sudoku80.txt",
                  "sudoku81.txt", "sudoku82.txt", "sudoku83.txt", "sudoku84.txt", "sudoku85.txt",
                  "sudoku86.txt", "sudoku87.txt", "sudoku88.txt", "sudoku89.txt", "sudoku90.txt",
                  "sudoku91.txt", "sudoku92.txt", "sudoku93.txt", "sudoku94.txt", "sudoku95.txt",
                  "sudoku96.txt", "sudoku97.txt", "sudoku98.txt", "sudoku99.txt", "sudoku100.txt"]
    
    zaehler = 1
    for dateipfad in dateipfade:
        matrix = datei_lesen(dateipfad)
        if matrix is not None:
            ausgabedatei = generiere_dateiname(f'matrix_{dateipfad}.xlsx')
            bearbeite_matrix_und_speichere_als_excel(matrix, ausgabedatei)
            print(f"Sudoku aus '{dateipfad}' erfolgreich konvertiert.\nIn '{ausgabedatei}' gespeichert.")

            # Excel-Datei mit pandas einlesen
            df = pd.read_excel(ausgabedatei, header=None, index_col=None)

            # HTML-Code der Tabelle generieren
            html_table = df.to_html(index=False, header=False, border=0, classes='sudoku-table')

            # HTML-Dokument erstellen und Tabelle mittig positionieren
            html_content = f"<html><head><style>.sudoku-table {{margin: 0 auto;}}</style></head><body>{html_table}"
            html_content += f'<p class="numbers">Nummer: {zaehler}</p>\n</body></html>'
            
   

            # HTML-Dokument speichern
            html_dateiname = generiere_dateiname(f'matrix_{dateipfad}.html')
            with open(html_dateiname, 'w') as html_datei:
                html_datei.write(html_content)
                
            print(f"HTML-Dokument mit Sudoku aus '{ausgabedatei}' erstellt.\nIn '{html_dateiname}' gespeichert.")
            
            zaehler += 1  # Erhöhe den Zähler für die nächste Iteration